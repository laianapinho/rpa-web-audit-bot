# Importa Path para trabalhar com caminhos de arquivos e criar pastas.
from pathlib import Path

# Importa pandas para transformar os resultados em DataFrame.
import pandas as pd

# Importa funções de estilo do openpyxl para formatar o Excel.
from openpyxl.styles import Font, PatternFill, Alignment

# Importa get_column_letter para ajustar largura das colunas.
from openpyxl.utils import get_column_letter

# Importa o caminho do arquivo de saída definido no config.py.
from app.config import OUTPUT_FILE

# Importa o logger do projeto.
from app.logger_config import configurar_logger


# Cria o logger deste arquivo.
logger = configurar_logger()


def criar_dataframe_auditoria(resultados_auditoria):
    # Converte a lista de dicionários em um DataFrame do Pandas.
    df = pd.DataFrame(resultados_auditoria)

    # Define a ordem das colunas no relatório.
    colunas = [
        "codigo",
        "nome_esperado",
        "nome_encontrado",
        "status_esperado",
        "status_encontrado",
        "resultado_auditoria",
        "mensagem",
        "evidencia"
    ]

    # Reorganiza o DataFrame seguindo a ordem definida.
    df = df[colunas]

    # Retorna o DataFrame pronto para exportação.
    return df

def calcular_percentual(valor, total):
    # Evita divisão por zero.
    if total == 0:
        return "0%"

    # Calcula o percentual.
    percentual = (valor / total) * 100

    # Retorna percentual formatado sem casas decimais.
    return f"{percentual:.0f}%"

def criar_dataframe_resumo(df_auditoria):
    # Calcula o total de registros processados.
    total_registros = len(df_auditoria)

    # Conta quantos registros ficaram como Conforme.
    total_conforme = len(
        df_auditoria[df_auditoria["resultado_auditoria"] == "Conforme"]
    )

    # Conta quantos registros ficaram como Não conforme.
    total_nao_conforme = len(
        df_auditoria[df_auditoria["resultado_auditoria"] == "Não conforme"]
    )

    # Conta quantos registros ficaram como Não encontrado.
    total_nao_encontrado = len(
        df_auditoria[df_auditoria["resultado_auditoria"] == "Não encontrado"]
    )
        
    # Cria uma tabela de resumo.
    dados_resumo = {
        "indicador": [
            "Total de registros",
            "Total conforme",
            "Total não conforme",
            "Total não encontrado"
        ],
        "valor": [
            total_registros,
            total_conforme,
            total_nao_conforme,
            total_nao_encontrado
        ],
        "percentual": [
            "100%",
            calcular_percentual(total_conforme, total_registros),
            calcular_percentual(total_nao_conforme, total_registros),
            calcular_percentual(total_nao_encontrado, total_registros)
        ]
    }

    # Converte o dicionário em DataFrame.
    df_resumo = pd.DataFrame(dados_resumo)

    # Retorna o resumo.
    return df_resumo


def ajustar_largura_colunas(worksheet):
    # Percorre todas as colunas da planilha.
    for coluna in worksheet.columns:
        # Define tamanho mínimo da coluna.
        maior_tamanho = 0

        # Pega a letra da coluna atual.
        letra_coluna = get_column_letter(coluna[0].column)

        # Percorre cada célula da coluna.
        for celula in coluna:
            # Se a célula tiver valor, calcula o tamanho do texto.
            if celula.value:
                tamanho_texto = len(str(celula.value))

                # Guarda o maior tamanho encontrado.
                if tamanho_texto > maior_tamanho:
                    maior_tamanho = tamanho_texto

        # Define a largura com uma pequena folga.
        largura_ajustada = maior_tamanho + 3

        # Limita largura máxima para evitar colunas gigantes.
        if largura_ajustada > 60:
            largura_ajustada = 60

        # Aplica a largura na coluna.
        worksheet.column_dimensions[letra_coluna].width = largura_ajustada


def formatar_cabecalho(worksheet):
    # Cria preenchimento azul para o cabeçalho.
    preenchimento_cabecalho = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    # Cria fonte branca e em negrito.
    fonte_cabecalho = Font(
        color="FFFFFF",
        bold=True
    )

    # Percorre as células da primeira linha.
    for celula in worksheet[1]:
        # Aplica cor de fundo.
        celula.fill = preenchimento_cabecalho

        # Aplica fonte.
        celula.font = fonte_cabecalho

        # Centraliza o texto.
        celula.alignment = Alignment(horizontal="center")


def formatar_resultados_auditoria(worksheet):
    # Define cores para cada resultado.
    cor_conforme = PatternFill(fill_type="solid", fgColor="C6EFCE")
    cor_nao_conforme = PatternFill(fill_type="solid", fgColor="FFC7CE")
    cor_nao_encontrado = PatternFill(fill_type="solid", fgColor="FFEB9C")

    # Identifica a coluna resultado_auditoria.
    coluna_resultado = None

    # Percorre o cabeçalho para achar a coluna.
    for celula in worksheet[1]:
        if celula.value == "resultado_auditoria":
            coluna_resultado = celula.column
            break

    # Se não encontrar a coluna, encerra a função.
    if coluna_resultado is None:
        return

    # Percorre as linhas de dados, começando da linha 2.
    for linha in range(2, worksheet.max_row + 1):
        # Pega a célula do resultado da auditoria.
        celula_resultado = worksheet.cell(
            row=linha,
            column=coluna_resultado
        )

        # Pega o valor da célula.
        valor = celula_resultado.value

        # Aplica cor conforme o resultado.
        if valor == "Conforme":
            celula_resultado.fill = cor_conforme

        elif valor == "Não conforme":
            celula_resultado.fill = cor_nao_conforme

        elif valor == "Não encontrado":
            celula_resultado.fill = cor_nao_encontrado

        # Centraliza o texto da célula.
        celula_resultado.alignment = Alignment(horizontal="center")


def formatar_planilha(worksheet):
    # Formata o cabeçalho.
    formatar_cabecalho(worksheet)

    # Ajusta largura das colunas.
    ajustar_largura_colunas(worksheet)

    # Congela a primeira linha.
    worksheet.freeze_panes = "A2"

    # Ativa filtro automático.
    worksheet.auto_filter.ref = worksheet.dimensions


def gerar_relatorio_excel(resultados_auditoria):
    # Registra no log que a geração do relatório começou.
    logger.info("Iniciando geração do relatório Excel.")

    # Garante que a pasta output existe.
    Path(OUTPUT_FILE).parent.mkdir(parents=True, exist_ok=True)

    # Cria o DataFrame principal da auditoria.
    df_auditoria = criar_dataframe_auditoria(resultados_auditoria)

    # Cria o DataFrame de resumo.
    df_resumo = criar_dataframe_resumo(df_auditoria)

    # Cria o arquivo Excel com duas abas.
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        # Salva a aba principal.
        df_auditoria.to_excel(
            writer,
            sheet_name="Auditoria",
            index=False
        )

        # Salva a aba de resumo.
        df_resumo.to_excel(
            writer,
            sheet_name="Resumo",
            index=False
        )

        # Acessa a aba Auditoria.
        aba_auditoria = writer.book["Auditoria"]

        # Acessa a aba Resumo.
        aba_resumo = writer.book["Resumo"]

        # Formata a aba Auditoria.
        formatar_planilha(aba_auditoria)

        # Aplica cores na coluna resultado_auditoria.
        formatar_resultados_auditoria(aba_auditoria)

        # Formata a aba Resumo.
        formatar_planilha(aba_resumo)

    # Registra no log o caminho do relatório.
    logger.info(f"Relatório Excel gerado com sucesso: {OUTPUT_FILE}")

    # Retorna o caminho do arquivo gerado.
    return OUTPUT_FILE